from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

filename="out/02-01-2020_11-42-14.xlsx"
filename2="out/02-01-2020_11-42-142.xlsx"

wb = Workbook()
wb = load_workbook(filename)
active = wb['sheet1']
wb.create_sheet('sheet2',1)
active2 = wb['sheet2']
newSheet=[]
for row in active:
    temp=[]
    temp.append(row[3])
    temp.append(row[0])
    temp.append(row[1])
    temp.append(row[2])
    newSheet.append(temp)

print (newSheet)
wb.save(filename2)
