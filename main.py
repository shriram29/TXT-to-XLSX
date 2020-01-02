from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os

def clear():
    if os.name == 'nt':
        _ = os.system('cls')
    else:
        _ = os.system('clear')

def printProgressBar (iteration, total, prefix = '', suffix = '', decimals = 1, length = 100, fill = '█', printEnd = "\r"):
    percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
    filledLength = int(length * iteration // total)
    bar = fill * filledLength + '-' * (length - filledLength)
    prefix="Line "+str(iteration)+" of "+str(total)
    print('\r%s |%s| %s%% %s' % (prefix, bar, percent, suffix), end = printEnd)
    if iteration == total:
        print()

def parse(x):
    out=[]
    for itm in x:
        if itm.isdigit() or (itm.startswith('-') and itm[1:].isdigit()):
            out.append(int(itm))
        else:
            out.append(itm.replace("_",":"))
    return out

now=datetime.now()
outFilename=now.strftime("./out/%d-%m-%Y_%H-%M-%S.xlsx")
loopvar=0
path = './in/'
fileNalmes = []
for r, d, f in os.walk(path):
    for file in f:
        if '.txt' in file:
            fileNalmes.append(file)

clear()
print("\n=== TXT to XLSX ===\n")
for file in fileNalmes:
    print(str(loopvar)+" | "+file)
    loopvar+=1
print("\nSelect File: ",end="")
intimp=int(input())
try:
    inputFilePath=path+fileNalmes[intimp]
    with open(inputFilePath) as inputFile:
        print("\n=== Deatils ===")
        print("\nOpened : "+inputFilePath[5:])
        for lineno, l in enumerate(inputFile):
            pass
        lineno=lineno+1
        print("\n"+str(lineno)+" lines in total\n")
    with open(inputFilePath) as inputFile:
        wb = Workbook()
        wb.create_sheet('sheet1',0)
        active = wb['sheet1']
        processStart=datetime.now()
        line = inputFile.readline()
        loopvar=1
        print("=== Processing ===\n")
        printProgressBar(loopvar, lineno, length = 50)
        while line:
           active.append(parse(line.split("\t|\t")))
           line = inputFile.readline()
           printProgressBar(loopvar, lineno,  length = 50)
           loopvar+=1
        processEnd=datetime.now()
        print("\n=== Output ===")
        print("\nProcessing Took : "+str(processEnd-processStart))
        wb.save(outFilename)
        print("\nFile Created :"+outFilename)
except Exception as error:
    print("\nError :" + repr(error)+" ")
print("\n=====================\n")
